#!/usr/bin/env python3
"""
Process job applications: organize files and update Excel tracker.
"""
import os
import sys
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font
from datetime import datetime
import re
import glob


def abbreviate_company(company):
    """Generate abbreviation for company name."""
    # Known company abbreviations
    known = {
        'equifax': 'EQF',
        'fgf brands': 'FGF',
        'geotab': 'GEO',
        'ibm': 'IBM',
        'nokia': 'NOK',
        'orsc': 'ORSC',
        'rbc': 'RBC',
        'rbcx': 'RBCx',
        'shoppers drug mart': 'SDM',
        'google': 'GOOG',
        'microsoft': 'MSFT',
        'amazon': 'AMZN',
        'apple': 'AAPL',
        'meta': 'META',
        'tesla': 'TSLA',
        'netflix': 'NFLX',
        'uber': 'UBER',
        'airbnb': 'ABNB',
        'spotify': 'SPOT',
        'shopify': 'SHOP',
        'td bank': 'TD',
        'bmo': 'BMO',
        'scotiabank': 'SCOT',
        'cibc': 'CIBC',
        'deloitte': 'DEL',
        'kpmg': 'KPMG',
        'pwc': 'PWC',
        'ey': 'EY',
        'accenture': 'ACN',
        'bnp paribas': 'BNPP',
    }

    company_lower = company.lower().strip()
    if company_lower in known:
        return known[company_lower]

    # If company name is already short (<=4 chars), use as-is
    if len(company) <= 4:
        return company.upper()

    # Generate abbreviation from first letters of words
    words = re.split(r'[\s\-]+', company)
    if len(words) >= 2:
        # Take first letter of each word (up to 4)
        abbr = ''.join(w[0].upper() for w in words[:4] if w)
        return abbr
    else:
        # Single word: take first 3-4 consonants or chars
        return company[:3].upper()


def abbreviate_position(position):
    """Generate abbreviation for position name."""
    # Common position word abbreviations
    word_abbr = {
        'analyst': 'A',
        'analytics': 'A',
        'assistant': 'A',
        'associate': 'A',
        'business': 'B',
        'co-op': '',
        'coop': '',
        'customer': 'C',
        'data': 'D',
        'developer': 'D',
        'delivery': 'D',
        'digital': 'D',
        'engineer': 'E',
        'engineering': 'E',
        'excellence': 'E',
        'financial': 'F',
        'global': 'G',
        'health': 'H',
        'insights': 'I',
        'intern': 'I',
        'internship': 'I',
        'it': 'IT',
        'machine': 'M',
        'marketing': 'M',
        'municipal': 'M',
        'learning': 'L',
        'operations': 'O',
        'policy': 'P',
        'product': 'P',
        'research': 'R',
        'revops': 'RO',
        'scientist': 'S',
        'services': 'S',
        'solutions': 'S',
        'systems': 'S',
        'software': 'SW',
        'telecom': 'T',
        'ai': 'AI',
        'ai-ml': 'AIML',
        'ml': 'ML',
    }

    # Extract parenthetical suffix like (RevOps), (Digital Health)
    suffix = ''
    paren_match = re.search(r'\(([^)]+)\)', position)
    if paren_match:
        paren_content = paren_match.group(1).lower()
        paren_words = re.split(r'[\s\-]+', paren_content)
        suffix_parts = []
        for w in paren_words:
            if w in word_abbr:
                if word_abbr[w]:
                    suffix_parts.append(word_abbr[w])
            else:
                suffix_parts.append(w[0].upper())
        if suffix_parts:
            suffix = '-' + ''.join(suffix_parts)
        position = re.sub(r'\s*\([^)]+\)', '', position)

    # Process main position
    words = re.split(r'[\s\-]+', position.lower())
    abbr_parts = []
    for w in words:
        if w in word_abbr:
            if word_abbr[w]:  # Skip empty abbreviations (co-op, coop)
                abbr_parts.append(word_abbr[w])
        elif w:
            abbr_parts.append(w[0].upper())

    return ''.join(abbr_parts) + suffix


def find_latest_files_in_exp_lib(exp_lib_dir):
    """
    Find R-*.pdf and CL-*.pdf files in exp_lib directory.

    Priority:
    1. Use RESUME_FILE and CL_FILE environment variables if set (for batch processing)
    2. Otherwise find the latest files by modification time

    Args:
        exp_lib_dir: Path to exp_lib directory

    Returns:
        Tuple of (resume_path, cover_letter_path) or (None, None) if not found
    """
    # Check for environment variables first (used by batch_organize.py)
    env_resume = os.environ.get('RESUME_FILE')
    env_cl = os.environ.get('CL_FILE')

    if env_resume or env_cl:
        resume_path = Path(env_resume) if env_resume and Path(env_resume).exists() else None
        cl_path = Path(env_cl) if env_cl and Path(env_cl).exists() else None

        if resume_path:
            print(f"📄 Using specified resume: {resume_path.name}")
        if cl_path:
            print(f"📄 Using specified cover letter: {cl_path.name}")

        return resume_path, cl_path

    # Fallback to finding latest files
    exp_lib_path = Path(exp_lib_dir)

    if not exp_lib_path.exists():
        print(f"⚠️  Warning: exp_lib directory not found at {exp_lib_dir}")
        return None, None

    # Find all R-*.pdf files
    resume_files = list(exp_lib_path.glob("R-*.pdf"))
    # Find all CL-*.pdf files
    cl_files = list(exp_lib_path.glob("CL-*.pdf"))

    # Get the latest files by modification time
    latest_resume = max(resume_files, key=lambda p: p.stat().st_mtime) if resume_files else None
    latest_cl = max(cl_files, key=lambda p: p.stat().st_mtime) if cl_files else None

    if latest_resume:
        print(f"📄 Found latest resume: {latest_resume.name}")
    else:
        print(f"⚠️  Warning: No R-*.pdf files found in {exp_lib_dir}")

    if latest_cl:
        print(f"📄 Found latest cover letter: {latest_cl.name}")
    else:
        print(f"⚠️  Warning: No CL-*.pdf files found in {exp_lib_dir}")

    return latest_resume, latest_cl


def backup_excel_file(excel_path):
    """
    Create a backup of the Excel file before modifying it.
    Backup format: job-applications_backup_YYYY-MM-DD_HH-MM-SS.xlsx
    """
    if not excel_path.exists():
        return None

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_dir = excel_path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)

    backup_path = backup_dir / f"job-applications_backup_{timestamp}.xlsx"
    shutil.copy2(excel_path, backup_path)
    print(f"📦 Created backup: {backup_path.name}")
    return backup_path


def find_excel_file(base_dir):
    """Find job-applications.xlsx in the project directory."""
    # Try common locations
    possible_paths = [
        Path(base_dir) / "job-applications.xlsx",
        Path(base_dir) / "job-applications" / "job-applications.xlsx",
    ]

    for path in possible_paths:
        if path.exists():
            return path

    # Search recursively
    for root, dirs, files in os.walk(base_dir):
        if "job-applications.xlsx" in files:
            return Path(root) / "job-applications.xlsx"

    return None


def get_next_row(ws):
    """Find the next available row in the Excel sheet."""
    # Start from row 2 (after header)
    row = 2
    while ws.cell(row, 1).value is not None:
        row += 1
    return row


def add_job_to_excel(excel_path, company, position, jd_text, folder_path):
    """Add a new job application to the Excel tracker."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Check if this is a new file (no headers)
    if ws.cell(1, 1).value is None:
        # Set up headers
        headers = ["Company", "Position", "Job Description", "Application Date", "Files"]
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col).value = header
            ws.cell(1, col).alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

        # Set column widths
        ws.column_dimensions['A'].width = 20  # Company
        ws.column_dimensions['B'].width = 30  # Position
        ws.column_dimensions['C'].width = 60  # Job Description
        ws.column_dimensions['D'].width = 15  # Application Date
        ws.column_dimensions['E'].width = 15  # Files

    # Get next available row
    row = get_next_row(ws)

    # Get current date
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Add data
    ws.cell(row, 1).value = company
    ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical='top')

    ws.cell(row, 2).value = position
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')

    # Store the COMPLETE ORIGINAL JD text unchanged
    ws.cell(row, 3).value = jd_text
    ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical='top')

    ws.cell(row, 4).value = current_date
    ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical='top')

    # Add hyperlink to folder
    ws.cell(row, 5).value = "Files"
    ws.cell(row, 5).hyperlink = str(folder_path)
    ws.cell(row, 5).font = Font(color="0563C1", underline="single")
    ws.cell(row, 5).alignment = Alignment(wrap_text=True, vertical='top')

    # Set row height to auto
    ws.row_dimensions[row].height = None

    wb.save(excel_path)
    print(f"✅ Added {company} - {position} to Excel tracker at row {row}")
    print(f"   JD text length: {len(jd_text)} characters (original text preserved)")


def organize_files(base_dir, company, position, resume_file, cl_file, application_date=None):
    """
    Organize resume and cover letter files by date.

    Args:
        base_dir: Base directory for job applications
        company: Company name
        position: Position name
        resume_file: Path to resume file (can be None)
        cl_file: Path to cover letter file (can be None)
        application_date: Application date (YYYY-MM-DD format), defaults to today

    Returns:
        Tuple of (folder_path, folder_reference_string)
    """
    # Use today's date if not provided
    if application_date is None:
        application_date = datetime.now().strftime("%Y-%m-%d")

    folder_name = f"{company} - {position}"
    # Organize by date: resume and cover letters/YYYY-MM-DD/Company - Position/
    job_dir = Path(base_dir) / "resume and cover letters" / application_date / folder_name
    job_dir.mkdir(parents=True, exist_ok=True)

    # Generate abbreviated prefix for filenames
    co_abbr = abbreviate_company(company)
    pos_abbr = abbreviate_position(position)
    prefix = f"{co_abbr} - {pos_abbr}"

    files_copied = []

    # Move resume file (not copy)
    if resume_file and Path(resume_file).exists():
        src = Path(resume_file)
        dst = job_dir / f"{prefix} - R.pdf"
        shutil.move(str(src), str(dst))
        print(f"✅ Organized resume: {dst}")
        print(f"   Removed from: {src}")
        files_copied.append("R")

    # Move cover letter file (not copy)
    if cl_file and Path(cl_file).exists():
        src = Path(cl_file)
        dst = job_dir / f"{prefix} - CL.pdf"
        shutil.move(str(src), str(dst))
        print(f"✅ Organized cover letter: {dst}")
        print(f"   Removed from: {src}")
        files_copied.append("CL")

    # Return the absolute path and a reference string
    folder_path = str(job_dir.absolute())
    if files_copied:
        return folder_path, f"See folder: {application_date}/{folder_name}/ ({', '.join(files_copied)})"
    else:
        return folder_path, f"See folder: {application_date}/{folder_name}/ (no files)"


def main():
    """Main entry point."""
    if len(sys.argv) < 5:
        print("Usage: process_job_application.py <base_dir> <company> <position> <jd_text> [exp_lib_dir]")
        print("  base_dir: Directory for job applications (default: /Users/samguan/Desktop/project/n8n-resume-generator/cvs)")
        print("  exp_lib_dir: Optional path to exp_lib directory (default: /Users/samguan/Desktop/project/n8n-resume-generator/exp_lib)")
        print("  jd_text: COMPLETE ORIGINAL JD TEXT (unchanged)")
        sys.exit(1)

    base_dir = sys.argv[1]
    company = sys.argv[2]
    position = sys.argv[3]
    jd_text = sys.argv[4]  # MUST be complete original JD text
    
    # Default exp_lib directory or use provided one
    if len(sys.argv) > 5:
        exp_lib_dir = sys.argv[5]
    else:
        exp_lib_dir = "/Users/samguan/Desktop/project/n8n-resume-generator/exp_lib"

    # Find Excel file
    excel_path = find_excel_file(base_dir)
    if not excel_path:
        print(f"❌ Error: Could not find job-applications.xlsx in {base_dir}")
        sys.exit(1)

    print(f"📊 Found Excel tracker: {excel_path}")

    # Backup Excel file before making changes
    backup_excel_file(excel_path)

    # Find latest resume and cover letter files in exp_lib
    print(f"\n🔍 Searching for files in: {exp_lib_dir}")
    resume_file, cl_file = find_latest_files_in_exp_lib(exp_lib_dir)
    
    if not resume_file and not cl_file:
        print(f"\n⚠️  Warning: No R-*.pdf or CL-*.pdf files found in {exp_lib_dir}")
        print("    Continuing without files...")

    # Get current date for organizing files
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Organize files and get folder path
    print(f"\n📁 Organizing files...")
    folder_path, folder_reference = organize_files(
        base_dir, company, position, resume_file, cl_file, current_date
    )

    # Add to Excel with COMPLETE ORIGINAL JD text
    print(f"\n📝 Updating Excel tracker...")
    add_job_to_excel(excel_path, company, position, jd_text, folder_path)

    print(f"\n✅ Successfully processed application for {company} - {position}")
    print(f"   Files organized in: resume and cover letters/{current_date}/{company} - {position}/")
    print(f"   Original JD text saved to Excel (unchanged)")


if __name__ == "__main__":
    main()
